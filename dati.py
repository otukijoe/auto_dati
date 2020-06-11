#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
'''
@File    :   dati.py
@Time    :   2020/06/10 08:47:57
@Author  :   otuki
@Version :   1.0
@Desc    :   otuki.top
'''

import re
import sys

import openpyxl


def load_xlsx(ws, istiku):
    '''
    @description: 载入题库或试题
    @param {type}
    @return: 
    '''
    tiku = []
    for row in ws.iter_rows():
        tk = {}
        try:
            tk["question"] = row[1].value.strip().replace("（", "(").replace(
                "）",
                ")").replace(" ",
                             "").replace("\xa0",
                                         "").replace("\n",
                                                     "").replace(" \t", "")
            tk["A"] = row[2].value.strip().replace("（", "(").replace(
                "）",
                ")").replace(" ",
                             "").replace("\xa0",
                                         "").replace("\n",
                                                     "").replace(" \t", "")
            tk["B"] = row[3].value.strip().replace("（", "(").replace(
                "）",
                ")").replace(" ",
                             "").replace("\xa0",
                                         "").replace("\n",
                                                     "").replace(" \t", "")
            tk["C"] = row[4].value.strip().replace("（", "(").replace(
                "）",
                ")").replace(" ",
                             "").replace("\xa0",
                                         "").replace("\n",
                                                     "").replace(" \t", "")
            tk["D"] = row[5].value.strip().replace("（", "(").replace(
                "）",
                ")").replace(" ",
                             "").replace("\xa0",
                                         "").replace("\n",
                                                     "").replace(" \t", "")
            if istiku:
                tk["answer"] = row[6].value.strip().replace(" ", "")
        except Exception as e:
            tk["question"] = row[1].value
            tk["A"] = row[2].value
            tk["B"] = row[3].value
            tk["C"] = row[4].value
            tk["D"] = row[5].value
            if istiku:
                tk["answer"] = row[6].value
        tiku.append(tk)
    return tiku


def load_txt(tiku_path, istiku):
    tiku = []
    with open(tiku_path, "r") as f:
        tk_list = f.readlines()
    tk = {}
    for tk_l in tk_list:
        try:
            tk_l = tk_l.strip().replace("（", "(").replace("）", ")").replace(
                " ", "").replace("\xa0", "").replace("\n",
                                                     "").replace(" \t", "")
        except Exception as e:
            tk_l = tk_l
        if tk_l[0] == "答":
            tk["answer"] = tk_l.split(':')[1]
        elif tk_l[0] == "A":
            tk["A"] = tk_l.split('.')[1]
        elif tk_l[0] == "B":
            tk["B"] = tk_l.split('.')[1]
        elif tk_l[0] == "C":
            tk["C"] = tk_l.split('.')[1]
        elif tk_l[0] == "D":
            tk["D"] = tk_l.split('.')[1]
        else:
            tiku.append(tk)
            tk = {}
            # 去掉"1."、"2."等题号
            reg = re.search(r'^([1-9]+\.)', tk_l).group(1)
            if reg != None:
                tk_l = tk_l.replace(reg, "")
            tk["question"] = tk_l
    tiku.append(tk)
    return tiku[1:]


def load_tiku(tiku_path):
    if tiku_path.endswith('.xlsx'):
        wb_tiku = openpyxl.load_workbook(tiku_path, read_only=True)
        ws_tiku = wb_tiku.active
        tiku = load_xlsx(ws_tiku, istiku=True)
    elif tiku_path.endswith('.txt'):
        tiku = load_txt(tiku_path, istiku=True)
    else:
        print("[!]题库文件格式错误！")
        sys.exit(0)
    print("[!]获取到题库!")
    # print(tiku)
    return tiku


def load_ti(ti_path):
    if ti_path.endswith('.xlsx'):
        wb_ti = openpyxl.load_workbook(ti_path, read_only=True)
        ws_ti = wb_ti.active
        ti = load_xlsx(ws_ti, istiku=False)
    elif ti_path.endswith('.txt'):
        ti = load_txt(ti_path, istiku=False)
    else:
        print("[!]题目文件格式错误！")
        sys.exit(0)
    print("[!]获取到题目!")
    # print(ti)
    return ti


def find_answer(tiku, ti):
    answers = []
    for i in range(len(ti)):
        ques = {}
        ques["num"] = i + 1
        ques['question'] = ti[i]['question']
        ans = []
        for tk in tiku:
            # 找到同一问题
            if ti[i]['question'] == tk['question']:
                # a in ["A","B","C","D"]
                for a in list(tk['answer']):
                    if tk[a] == ti[i]["A"]:
                        ans.append({"A": tk[a]})
                    elif tk[a] == ti[i]["B"]:
                        ans.append({"B": tk[a]})
                    elif tk[a] == ti[i]["C"]:
                        ans.append({"C": tk[a]})
                    elif tk[a] == ti[i]["D"]:
                        ans.append({"D": tk[a]})
            # 得到答案
            if len(ans) > 0:
                ques["answers"] = ans
                break
        if len(ans) == 0:
            ques["answers"] = None
        answers.append(ques)
    return answers


def write_xlsx(ti_path, find):
    wb_ti = openpyxl.load_workbook(ti_path, read_only=False)
    ws_ti = wb_ti.active
    for f in find:
        v = ''
        for a in f["answers"]:
            v += list(a.keys())[0]
        # print(f["num"], v)
        ws_ti.cell(row=f["num"], column=7, value=v)
    wb_ti.save(ti_path)


def start(tiku_path, ti_path):
    tiku = load_tiku(tiku_path)
    ti = load_ti(ti_path)
    answers = find_answer(tiku, ti)
    find = []
    notfind = []
    for ques in answers:
        if ques["answers"] != None:
            find.append(ques)
        else:
            notfind.append(ques)
    print("[!]已找到答案:")
    for f in find:
        print(f['num'], end=' ')
    # print("\n", find)
    if len(notfind) > 0:
        print("\n[!]未找到答案:")
        for n in notfind:
            print(n['num'], end=' ')
    if ti_path.endswith('.xlsx'):
        write_xlsx(ti_path, find)
        print("\n--------------------------------")
        print("[!]答案已写入xlsx.")
    else:
        print("\n--------------------------------")
        for f in find:
            v = ''
            for a in f["answers"]:
                v += list(a.keys())[0]
            print(f["num"], ": " + v)


if __name__ == "__main__":
    print("--------------------------------")
    print("-------------自动答题------------")
    print("--------------------------------")
    print("----------Author:otuki----------")
    print("--------------------------------")
    tiku_path = "/Users/joe/Desktop/自动化竞赛/3.竞赛题库/综合.xlsx"
    ti_path = "/Users/joe/Desktop/1.txt"
    start(tiku_path, ti_path)
